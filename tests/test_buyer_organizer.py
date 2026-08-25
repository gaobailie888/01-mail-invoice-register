import json
import sys
import tempfile
import unittest
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))
import buyer_organizer as buyer


class BuyerOrganizerTests(unittest.TestCase):
    def row(self, number, digest, amount=100, name="购方甲", filename="a.pdf"):
        return {"发票号码": number, "SHA256": digest, "价税合计": amount, "合计金额": amount - 6, "合计税额": 6, "购方名称": name, "附件文件名": filename, "本地文件": ""}

    def test_empty_numbers_are_not_deduplicated(self):
        rows = buyer.enrich_rows([self.row("", "h1"), self.row("", "h2")], {"file_to_buyer": {}, "buyer_aliases": {}})
        self.assertTrue(all(r["计入建议合计"] == "是" for r in rows))
        self.assertTrue(all("无法自动去重" in r["重复状态"] for r in rows))

    def test_exact_hash_duplicate_excludes_only_later_copy(self):
        rows = buyer.enrich_rows([self.row("10000001", "same"), self.row("10000001", "same")], {"file_to_buyer": {}, "buyer_aliases": {}})
        self.assertEqual([r["计入建议合计"] for r in rows], ["是", "否"])

    def test_same_number_different_content_is_kept(self):
        rows = buyer.enrich_rows([self.row("10000001", "h1"), self.row("10000001", "h2")], {"file_to_buyer": {}, "buyer_aliases": {}})
        self.assertTrue(all(r["计入建议合计"] == "是" for r in rows))
        self.assertTrue(all("内容不同" in r["重复状态"] for r in rows))

    def test_external_rules_and_alias(self):
        rules = {"file_to_buyer": {"x.pdf": "别名公司"}, "buyer_aliases": {"别名公司": "标准公司"}}
        rows = buyer.enrich_rows([self.row("10000001", "h", name="", filename="x.pdf")], rules)
        self.assertEqual(rows[0]["购方分类"], "标准公司")
        self.assertEqual(rows[0]["购方分类来源"], "用户规则")

    def test_unique_sheet_name(self):
        used = set()
        a = buyer.unique_sheet_name("A" * 40, used)
        b = buyer.unique_sheet_name("A" * 39 + "B", used)
        self.assertNotEqual(a, b)
        self.assertLessEqual(len(b), 31)

    def test_workbook_output_and_copy(self):
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            attachment = root / "a.pdf"
            attachment.write_bytes(b"same")
            wb = openpyxl.Workbook()
            summary = wb.active
            summary.title = "发票汇总"
            headers = ["发票号码","开票日期","销方名称","购方名称","价税合计","合计金额","合计税额","来源发件人","邮件日期","邮件主题","附件文件名","本地文件","解析状态","校验结果"]
            summary.append(headers)
            summary.append(["10000001","2026-08-01","销方","购方甲",106,100,6,"a","d","s","a.pdf",str(attachment),"已解析","通过"])
            log = wb.create_sheet("处理日志")
            log.append(["邮件ID","发件人","邮件日期","邮件主题","附件文件名","文件类型","本地文件","SHA256","状态","说明"])
            log.append(["1","a","d","s","a.pdf",".pdf",str(attachment),buyer.sha256_file(attachment),"已解析","通过"])
            wb.create_sheet("项目明细")
            source = root / "base.xlsx"
            wb.save(source)
            result = buyer.organize(source, root, root / "out")
            self.assertTrue(Path(result["buyer_workbook"]["output_file"]).is_file())
            self.assertTrue((root / "out" / "by_company" / "购方甲" / "a.pdf").is_file())
            out = openpyxl.load_workbook(result["buyer_workbook"]["output_file"])
            self.assertIn("购方索引", out.sheetnames)
            self.assertIn("统计口径", out.sheetnames)


if __name__ == "__main__":
    unittest.main()
