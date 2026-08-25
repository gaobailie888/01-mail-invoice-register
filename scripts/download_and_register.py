#!/usr/bin/env python3
"""Download invoice-related IMAP attachments and create an auditable workbook."""

from __future__ import annotations

import argparse
import email
import hashlib
import imaplib
import json
import re
import shutil
import sys
from email import policy
from collections import Counter
from dataclasses import asdict, dataclass, field
from datetime import date, datetime, timedelta
from email.header import decode_header
from email.message import Message
from email.utils import parsedate_to_datetime
from pathlib import Path
from typing import Iterable

import openpyxl
import pdfplumber
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


DEFAULT_KEYWORDS = ("发票", "电子发票", "增值税", "invoice", "receipt", "tax")
DOWNLOADABLE_EXTENSIONS = {".pdf", ".ofd", ".xml", ".jpg", ".jpeg", ".png", ".tif", ".tiff"}


@dataclass
class MailAttachment:
    message_id: str
    sender: str
    subject: str
    mail_date: str
    filename: str
    suffix: str
    local_path: str = ""
    sha256: str = ""
    status: str = "待下载"
    note: str = ""
    invoice: dict = field(default_factory=dict)


def validate_config(values: dict[str, str]) -> dict[str, str]:
    """Validate IMAP settings without logging or returning the password."""
    values = {key: str(value).strip() for key, value in values.items()}
    missing = [key for key in ("IMAP_HOST", "IMAP_USERNAME", "IMAP_PASSWORD") if not values.get(key)]
    if missing:
        raise ValueError("配置缺少：" + ", ".join(missing))
    if any(char.isspace() for char in values["IMAP_HOST"]):
        raise ValueError("IMAP_HOST 不能包含空格")
    port = int(values.get("IMAP_PORT", "993"))
    if not 1 <= port <= 65535:
        raise ValueError("IMAP_PORT 必须在 1 到 65535 之间")
    values["IMAP_PORT"] = str(port)
    values.setdefault("IMAP_MAILBOX", "INBOX")
    values.setdefault("IMAP_USE_SSL", "true")
    return values


def parse_env(path: Path) -> dict[str, str]:
    values: dict[str, str] = {}
    for raw in path.read_text(encoding="utf-8-sig").splitlines():
        line = raw.strip()
        if line and not line.startswith("#") and "=" in line:
            key, value = line.split("=", 1)
            values[key.strip()] = value.strip().strip('"').strip("'")
    return validate_config(values)


def decode_header_value(value: str | None) -> str:
    if not value:
        return ""
    result = []
    for fragment, charset in decode_header(value):
        result.append(fragment.decode(charset or "utf-8", errors="replace") if isinstance(fragment, bytes) else fragment)
    return "".join(result)


def safe_filename(name: str, fallback: str) -> str:
    name = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", name).strip(". ")
    return name[:160] or fallback


def message_text(message: Message) -> str:
    text: list[str] = []
    for part in message.walk():
        if part.get_content_maintype() == "multipart" or part.get_filename() or part.get_content_type() not in {"text/plain", "text/html"}:
            continue
        payload = part.get_payload(decode=True)
        if payload:
            text.append(payload.decode(part.get_content_charset() or "utf-8", errors="replace"))
    return "\n".join(text)[:20000]


def attachment_parts(message: Message) -> Iterable[Message]:
    """Return attachments for both legacy Message and EmailMessage objects."""
    for part in message.walk():
        if part.is_multipart():
            continue
        if part.get_filename() or part.get_content_disposition() == "attachment":
            yield part


def iter_candidates(message: Message, uid: str, keywords: tuple[str, ...], include_all_pdfs: bool, allowed_extensions: set[str]) -> Iterable[tuple[MailAttachment, bytes]]:
    subject = decode_header_value(message.get("Subject"))
    sender = decode_header_value(message.get("From"))
    try:
        mail_date = parsedate_to_datetime(message.get("Date", "")).astimezone().isoformat(timespec="seconds")
    except (TypeError, ValueError, OverflowError):
        mail_date = message.get("Date", "")
    body = (subject + "\n" + message_text(message)).lower()
    for index, part in enumerate(attachment_parts(message), 1):
        filename = decode_header_value(part.get_filename())
        suffix = Path(filename).suffix.lower()
        if suffix not in allowed_extensions:
            continue
        relevant = include_all_pdfs and suffix == ".pdf"
        relevant = relevant or any(word.lower() in (body + "\n" + filename.lower()) for word in keywords)
        payload = part.get_payload(decode=True)
        if not relevant or not payload:
            continue
        yield MailAttachment(uid, sender, subject, mail_date, safe_filename(filename, f"attachment-{index}{suffix}"), suffix), payload


def parse_amount(value: str | None) -> float | None:
    try:
        return float((value or "").replace(",", ""))
    except ValueError:
        return None


def find_text(patterns: Iterable[str], text: str) -> str:
    for pattern in patterns:
        match = re.search(pattern, text, re.S)
        if match:
            return re.sub(r"\s+", " ", match.group(1)).strip(" ：:")
    return ""


def parse_item(text: str) -> dict | None:
    values = re.findall(r"-?[\d,]+\.\d{2}", text)
    first_value = re.search(r"-?[\d,]+\.\d{2}", text)
    if len(values) < 2 or not first_value:
        return None
    return {"项目名称": re.sub(r"\s+", " ", text[:first_value.start()]).strip(), "金额": parse_amount(values[-2]), "税额": parse_amount(values[-1])}


def parse_items(lines: list[str]) -> list[dict]:
    items, current = [], []
    for raw in lines:
        line = raw.strip()
        if "价税合计" in line:
            break
        if line.startswith("*"):
            if current and (item := parse_item(" ".join(current))):
                items.append(item)
            current = [line]
        elif current and line:
            current.append(line)
    if current and (item := parse_item(" ".join(current))):
        items.append(item)
    return items


def validate_invoice(invoice: dict) -> None:
    missing = [key for key in ("发票号码", "销方名称", "价税合计") if not invoice.get(key)]
    if missing:
        invoice["解析状态"] = "待人工处理"
        invoice["校验结果"] = "缺少：" + "、".join(missing)
        return
    if all(invoice.get(key) is not None for key in ("价税合计", "合计金额", "合计税额")):
        delta = abs(invoice["价税合计"] - invoice["合计金额"] - invoice["合计税额"])
        invoice["校验结果"] = "通过" if delta <= 0.02 else f"价税合计差异 {delta:.2f}"
    else:
        invoice["校验结果"] = "缺少合计金额或税额，待核对"


def parse_invoice_pdf(path: Path) -> dict:
    invoice = {"发票号码": "", "开票日期": "", "销方名称": "", "购方名称": "", "价税合计": None, "合计金额": None, "合计税额": None, "项目明细": [], "解析状态": "已解析", "校验结果": "待校验"}
    try:
        with pdfplumber.open(path) as pdf:
            text = "\n".join(page.extract_text() or "" for page in pdf.pages)
    except Exception as exc:
        invoice.update({"解析状态": "待人工处理", "校验结果": f"PDF无法读取：{type(exc).__name__}"})
        return invoice
    if len(re.sub(r"\s+", "", text)) < 30:
        invoice.update({"解析状态": "待人工处理", "校验结果": "扫描件或无可提取文本"})
        return invoice
    invoice["发票号码"] = find_text((r"发票号码\s*[：:]?\s*(\d{8,22})",), text)
    invoice["开票日期"] = find_text((r"开票日期\s*[：:]?\s*(20\d{2}[年\-/]\d{1,2}[月\-/]\d{1,2}日?)",), text)
    invoice["销方名称"] = find_text((r"销售方\s*(?:名称)?\s*[：:]?\s*([^\n]{2,80}?)(?=\s*(?:统一社会信用|纳税人识别|地址|开户行|$))", r"销\s*售\s*方[\s\S]{0,100}?名称\s*[：:]?\s*([^\n]{2,80}?)(?=\s*(?:统一社会信用|纳税人识别|地址|开户行|$))", r"销\s*名称\s*[：:]?\s*([^\n]{2,80}?)(?=\s*(?:统一社会信用|纳税人识别|地址|开户行|$))"), text)
    invoice["购方名称"] = find_text((r"购买方\s*(?:名称)?\s*[：:]?\s*([^\n]{2,80}?)(?=\s*(?:统一社会信用|纳税人识别|地址|开户行|$))", r"购\s*买\s*方[\s\S]{0,100}?名称\s*[：:]?\s*([^\n]{2,80}?)(?=\s*(?:统一社会信用|纳税人识别|地址|开户行|$))", r"购\s*名称\s*[：:]?\s*([^\n]{2,80}?)(?=\s*(?:统一社会信用|纳税人识别|地址|开户行|$))"), text)
    for line in text.splitlines():
        names = [name.strip() for name in re.findall(r"名称[：:]\s*(.*?)(?=\s+名称[：:]|$)", line) if name.strip()]
        if len(names) >= 2:
            invoice["购方名称"] = invoice["购方名称"] or names[0]
            invoice["销方名称"] = invoice["销方名称"] or names[1]
            break
    invoice["价税合计"] = parse_amount(find_text((r"价\s*税\s*合\s*计[\s\S]{0,100}?[¥￥]?\s*(-?[\d,]+\.\d{2})",), text))
    totals = re.search(r"合\s*计\s*[¥￥]?\s*(-?[\d,]+\.\d{2})\s*[¥￥]?\s*(-?[\d,]+\.\d{2})", text.replace("\n", " "))
    if totals:
        invoice["合计金额"], invoice["合计税额"] = parse_amount(totals.group(1)), parse_amount(totals.group(2))
    invoice["项目明细"] = parse_items(text.splitlines())
    validate_invoice(invoice)
    return invoice


def apply_duplicate_flags(sheet) -> None:
    counts = Counter(sheet.cell(row, 1).value for row in range(2, sheet.max_row + 1) if sheet.cell(row, 1).value)
    for row in range(2, sheet.max_row + 1):
        number = sheet.cell(row, 1).value
        if number and counts[number] > 1:
            cell = sheet.cell(row, 14)
            cell.value = f"{cell.value or ''}；重复发票号码".strip("；")
            cell.fill = PatternFill("solid", fgColor="FFF2CC")


def create_workbook(records: list[MailAttachment], output_path: Path) -> None:
    workbook = openpyxl.Workbook()
    summary, detail, log = workbook.active, workbook.create_sheet("项目明细"), workbook.create_sheet("处理日志")
    summary.title = "发票汇总"
    summary_headers = ["发票号码", "开票日期", "销方名称", "购方名称", "价税合计", "合计金额", "合计税额", "来源发件人", "邮件日期", "邮件主题", "附件文件名", "本地文件", "解析状态", "校验结果"]
    detail_headers = ["发票号码", "销方名称", "项目名称", "金额", "税额", "附件文件名"]
    log_headers = ["邮件ID", "发件人", "邮件日期", "邮件主题", "附件文件名", "文件类型", "本地文件", "SHA256", "状态", "说明"]
    for sheet, headers in ((summary, summary_headers), (detail, detail_headers), (log, log_headers)):
        sheet.append(headers)
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.font, cell.fill, cell.alignment = Font(bold=True, color="FFFFFF"), PatternFill("solid", fgColor="1F4E78"), Alignment(horizontal="center")
    for record in records:
        invoice = record.invoice
        summary.append([invoice.get(key) for key in summary_headers[:7]] + [record.sender, record.mail_date, record.subject, record.filename, record.local_path, invoice.get("解析状态", record.status), invoice.get("校验结果", record.note)])
        if record.invoice:
            for item in invoice.get("项目明细", []):
                detail.append([invoice.get("发票号码"), invoice.get("销方名称"), item.get("项目名称"), item.get("金额"), item.get("税额"), record.filename])
        log.append([record.message_id, record.sender, record.mail_date, record.subject, record.filename, record.suffix, record.local_path, record.sha256, record.status, record.note])
    apply_duplicate_flags(summary)
    for row in range(2, summary.max_row + 1):
        invoice_number = summary.cell(row, 1)
        invoice_number.number_format = "@"
        if invoice_number.value is not None:
            invoice_number.value = str(invoice_number.value)
    border = Side(style="thin", color="D9E2F3")
    for sheet in workbook.worksheets:
        sheet.auto_filter.ref = sheet.dimensions
        for row in sheet.iter_rows():
            for cell in row:
                cell.border = Border(left=border, right=border, top=border, bottom=border)
                if isinstance(cell.value, float):
                    cell.number_format = '#,##0.00'
        for column in range(1, sheet.max_column + 1):
            longest = max((len(str(sheet.cell(row, column).value or "")) for row in range(1, sheet.max_row + 1)), default=0)
            sheet.column_dimensions[get_column_letter(column)].width = min(max(12, longest + 2), 45)
    workbook.save(output_path)


def imap_date(value: date) -> str:
    return value.strftime("%d-%b-%Y")


def open_mailbox(config: dict[str, str]):
    config = validate_config(config)
    use_ssl = config.get("IMAP_USE_SSL", "true").lower() not in {"0", "false", "no"}
    client = imaplib.IMAP4_SSL(config["IMAP_HOST"], int(config["IMAP_PORT"])) if use_ssl else imaplib.IMAP4(config["IMAP_HOST"], int(config.get("IMAP_PORT", "143")))
    client.login(config["IMAP_USERNAME"], config["IMAP_PASSWORD"])
    if client.select(config.get("IMAP_MAILBOX", "INBOX"), readonly=True)[0] != "OK":
        try:
            client.logout()
        finally:
            raise RuntimeError("无法以只读方式打开邮箱文件夹")
    return client


def test_connection(config: dict[str, str]) -> dict:
    client = open_mailbox(config)
    try:
        return {"ok": True, "mailbox": config.get("IMAP_MAILBOX", "INBOX"), "message": "连接成功，邮箱以只读方式打开。"}
    finally:
        try:
            client.logout()
        except Exception:
            pass


def process_local_folder(input_dir: Path, output_dir: Path, pdf_only: bool = False) -> list[MailAttachment]:
    """Copy and parse invoice-like files from a user-selected local folder."""
    input_dir = Path(input_dir).expanduser().resolve()
    output_dir = Path(output_dir).expanduser().resolve()
    if not input_dir.is_dir():
        raise ValueError("找不到发票输入文件夹")
    attachments_dir = output_dir / "attachments"
    attachments_dir.mkdir(parents=True, exist_ok=True)
    allowed = {".pdf"} if pdf_only else DOWNLOADABLE_EXTENSIONS
    records: list[MailAttachment] = []
    for source in sorted(path for path in input_dir.iterdir() if path.is_file() and path.suffix.lower() in allowed):
        target = attachments_dir / safe_filename(source.name, f"invoice{source.suffix.lower()}")
        version = 1
        while target.exists():
            target = attachments_dir / f"{source.stem}_{version}{source.suffix.lower()}"
            version += 1
        shutil.copy2(source, target)
        payload = target.read_bytes()
        record = MailAttachment("local", "本地文件夹", input_dir.name, datetime.fromtimestamp(source.stat().st_mtime).isoformat(timespec="seconds"), source.name, source.suffix.lower(), str(target), hashlib.sha256(payload).hexdigest())
        if record.suffix == ".pdf":
            record.invoice = parse_invoice_pdf(target)
            record.status, record.note = record.invoice["解析状态"], record.invoice["校验结果"]
        else:
            record.status, record.note = "待人工处理", f"{record.suffix or '未知'} 文件已复制，当前仅自动解析 PDF"
        records.append(record)
    return records


def run_local_folder_job(input_dir: Path, output_dir: Path, pdf_only: bool = False) -> tuple[list[MailAttachment], dict]:
    output_dir = Path(output_dir).expanduser().resolve()
    output_dir.mkdir(parents=True, exist_ok=True)
    records = process_local_folder(input_dir, output_dir, pdf_only)
    output = output_dir / f"AI本地发票登记{date.today():%Y%m%d}.xlsx"
    create_workbook(records, output)
    summary = build_summary(records, str(output))
    summary["source_mode"] = "local_folder"
    (output_dir / "run-summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    return records, summary


def build_summary(records: list[MailAttachment], output_file: str = "") -> dict:
    invoice_numbers = [item.invoice.get("发票号码") for item in records if item.invoice.get("发票号码")]
    counts = Counter(invoice_numbers)
    return {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "candidates": len(records),
        "pdf_files": sum(item.suffix == ".pdf" for item in records),
        "parsed_success": sum(item.invoice.get("解析状态") == "已解析" for item in records),
        "manual_review": sum(item.status == "待人工处理" or item.invoice.get("解析状态") == "待人工处理" for item in records),
        "duplicate_invoice_numbers": sum(count - 1 for count in counts.values() if count > 1),
        "output_file": output_file,
    }


def run_job(config: dict[str, str], since: date, until: date, output_dir: Path, keywords: tuple[str, ...] = DEFAULT_KEYWORDS, include_all_pdfs: bool = False, pdf_only: bool = False, dry_run: bool = False) -> tuple[list[MailAttachment], dict]:
    if until < since:
        raise ValueError("结束日期不能早于开始日期")
    output_dir = Path(output_dir).expanduser().resolve()
    output_dir.mkdir(parents=True, exist_ok=True)
    allowed_extensions = {".pdf"} if pdf_only else DOWNLOADABLE_EXTENSIONS
    records = download_records(validate_config(config), since, until, keywords, include_all_pdfs, output_dir / "attachments", dry_run, allowed_extensions)
    if dry_run:
        return records, build_summary(records)
    output = output_dir / f"AI邮箱发票登记{date.today():%Y%m%d}.xlsx"
    create_workbook(records, output)
    summary = build_summary(records, str(output))
    (output_dir / "run-summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    return records, summary


def download_records(config: dict[str, str], since: date, until: date, keywords: tuple[str, ...], include_all_pdfs: bool, attachments_dir: Path, dry_run: bool, allowed_extensions: set[str] = DOWNLOADABLE_EXTENSIONS) -> list[MailAttachment]:
    client = open_mailbox(config)
    records: list[MailAttachment] = []
    try:
        status, result = client.uid("SEARCH", None, "SINCE", imap_date(since), "BEFORE", imap_date(until + timedelta(days=1)))
        if status != "OK":
            raise RuntimeError("邮箱日期检索失败")
        if not dry_run:
            attachments_dir.mkdir(parents=True, exist_ok=True)
        for raw_uid in result[0].split():
            uid = raw_uid.decode()
            status, fetched = client.uid("FETCH", raw_uid, "(RFC822)")
            if status != "OK" or not fetched or not fetched[0]:
                continue
            message = email.message_from_bytes(fetched[0][1], policy=policy.default)
            for record, payload in iter_candidates(message, uid, keywords, include_all_pdfs, allowed_extensions):
                if dry_run:
                    record.status = "预览：未下载"
                    records.append(record)
                    continue
                target = attachments_dir / f"{uid}_{record.filename}"
                suffix, version = target.suffix, 1
                while target.exists():
                    target = attachments_dir / f"{uid}_{target.stem}_{version}{suffix}"
                    version += 1
                target.write_bytes(payload)
                record.local_path, record.sha256 = str(target), hashlib.sha256(payload).hexdigest()
                if record.suffix == ".pdf":
                    record.invoice = parse_invoice_pdf(target)
                    record.status, record.note = record.invoice["解析状态"], record.invoice["校验结果"]
                else:
                    record.status, record.note = "待人工处理", f"{record.suffix or '未知'} 文件已下载，当前仅自动解析 PDF"
                records.append(record)
    finally:
        try:
            client.logout()
        except Exception:
            pass
    return records


def main() -> int:
    parser = argparse.ArgumentParser(description="从 IMAP 邮箱下载发票附件并登记到 Excel")
    parser.add_argument("--config", type=Path, help="IMAP .env 配置文件；邮箱模式必填")
    parser.add_argument("--input-dir", type=Path, help="改用本地文件夹中的发票附件，不连接邮箱")
    parser.add_argument("--since", type=date.fromisoformat, default=date.today() - timedelta(days=30))
    parser.add_argument("--until", type=date.fromisoformat, default=date.today())
    parser.add_argument("--output-dir", type=Path, required=True)
    parser.add_argument("--keywords", default=",".join(DEFAULT_KEYWORDS))
    parser.add_argument("--include-all-pdfs", action="store_true")
    parser.add_argument("--pdf-only", action="store_true", help="仅下载和处理 PDF 附件")
    parser.add_argument("--dry-run", action="store_true", help="仅列出候选附件，不下载也不生成表格")
    args = parser.parse_args()
    if args.until < args.since:
        parser.error("--until 不能早于 --since")
    if args.input_dir:
        if args.dry_run:
            parser.error("本地文件夹模式不需要 --dry-run；只会复制文件，不会改动原文件")
        records, summary = run_local_folder_job(args.input_dir, args.output_dir, args.pdf_only)
    else:
        if not args.config or not args.config.is_file():
            parser.error("邮箱模式需要有效的 --config；或使用 --input-dir 处理本地文件夹")
        records, summary = run_job(
            parse_env(args.config), args.since, args.until, args.output_dir,
            tuple(item.strip() for item in args.keywords.split(",") if item.strip()),
            args.include_all_pdfs, args.pdf_only, args.dry_run,
        )
        if args.dry_run:
            summary["items"] = [asdict(item) for item in records]
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    sys.exit(main())
