#!/usr/bin/env python3
"""Organize invoice files by buyer and create safe buyer summary sheets."""
from __future__ import annotations
import argparse
import hashlib
import json
import re
import shutil
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

UNCLASSIFIED = "00_未分类_需人工确认"
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
NOTE_FILL = PatternFill("solid", fgColor="FFF2CC")
DUP_FILL = PatternFill("solid", fgColor="FCE4D6")
TOTAL_FILL = PatternFill("solid", fgColor="DDEBF7")
THIN = Side(style="thin", color="D9E2F3")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
VALID_NUMBER = re.compile(r"^\d{8,22}$")


def load_rules(path: Path | None) -> dict:
    if not path:
        return {"file_to_buyer": {}, "buyer_aliases": {}}
    data = json.loads(Path(path).read_text(encoding="utf-8"))
    return {
        "file_to_buyer": {str(k): str(v) for k, v in data.get("file_to_buyer", {}).items()},
        "buyer_aliases": {str(k): str(v) for k, v in data.get("buyer_aliases", {}).items()},
    }


def safe_dirname(name: str) -> str:
    cleaned = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", str(name)).strip(" .")
    return cleaned[:120] or UNCLASSIFIED


def clean_buyer(name, aliases: dict[str, str] | None = None) -> str:
    value = re.sub(r"\s+", " ", str(name or "")).strip()
    value = re.sub(r"\s*[销售购买方]{1,3}\s*$", "", value).strip()
    value = re.sub(r"\s+[销售]\s*$", "", value).strip()
    return (aliases or {}).get(value, value)


def unique_sheet_name(buyer: str, used: set[str]) -> str:
    base = re.sub(r"[\\/*?:\[\]]", "_", "购方-" + buyer).strip() or "购方-未分类"
    base = base[:31]
    name, index = base, 2
    while name in used:
        suffix = f"-{index}"
        name = base[:31 - len(suffix)] + suffix
        index += 1
    used.add(name)
    return name


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def enrich_rows(rows: list[dict], rules: dict) -> list[dict]:
    aliases, file_rules = rules["buyer_aliases"], rules["file_to_buyer"]
    number_counts = Counter(str(row.get("发票号码") or "").strip() for row in rows if VALID_NUMBER.fullmatch(str(row.get("发票号码") or "").strip()))
    hash_counts = Counter(str(row.get("SHA256") or "").strip() for row in rows if str(row.get("SHA256") or "").strip())
    seen_hashes: set[str] = set()
    result = []
    for row in rows:
        item = dict(row)
        filename = str(item.get("附件文件名") or "")
        buyer = clean_buyer(item.get("购方名称"), aliases)
        source = "发票解析"
        if not buyer and filename in file_rules:
            buyer, source = clean_buyer(file_rules[filename], aliases), "用户规则"
        if not buyer:
            buyer, source = "未分类", "待人工确认"
        number = str(item.get("发票号码") or "").strip()
        digest = str(item.get("SHA256") or "").strip()
        if digest and hash_counts[digest] > 1:
            if digest in seen_hashes:
                duplicate = "文件完全重复副本"
                include = False
            else:
                duplicate = "文件哈希重复，保留首份"
                include = True
                seen_hashes.add(digest)
        elif not VALID_NUMBER.fullmatch(number):
            duplicate = "无有效发票号码，无法自动去重"
            include = True
        elif number_counts[number] > 1:
            duplicate = "发票号码重复但内容不同，需人工确认"
            include = True
        else:
            duplicate = "未发现重复"
            include = True
        item.update({"购方分类": buyer, "购方分类来源": source, "重复状态": duplicate, "计入建议合计": "是" if include else "否"})
        result.append(item)
    return result


def rows_from_workbook(workbook_path: Path) -> tuple[list[str], list[dict]]:
    wb = openpyxl.load_workbook(workbook_path, data_only=False)
    summary = wb["发票汇总"]
    headers = [str(cell.value or "") for cell in summary[1]]
    log_hashes = {}
    if "处理日志" in wb.sheetnames:
        log = wb["处理日志"]
        log_headers = [str(cell.value or "") for cell in log[1]]
        name_idx = log_headers.index("附件文件名") if "附件文件名" in log_headers else -1
        hash_idx = log_headers.index("SHA256") if "SHA256" in log_headers else -1
        if name_idx >= 0 and hash_idx >= 0:
            for values in log.iter_rows(min_row=2, values_only=True):
                if values[name_idx]:
                    log_hashes[str(values[name_idx])] = str(values[hash_idx] or "")
    rows = []
    for values in summary.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, values))
        row["SHA256"] = log_hashes.get(str(row.get("附件文件名") or ""), "")
        rows.append(row)
    return headers, rows


def copy_by_buyer(rows: list[dict], output_dir: Path) -> dict:
    copied = skipped = missing = 0
    counts = Counter()
    for row in rows:
        source = Path(str(row.get("本地文件") or ""))
        buyer = str(row["购方分类"])
        counts[buyer] += 1
        folder = output_dir / safe_dirname(buyer if buyer != "未分类" else UNCLASSIFIED)
        folder.mkdir(parents=True, exist_ok=True)
        if not source.is_file():
            missing += 1
            continue
        target = folder / source.name
        if target.exists():
            skipped += 1
        else:
            shutil.copy2(source, target)
            copied += 1
    return {"copied": copied, "skipped_existing": skipped, "missing_source": missing, "buyer_counts": dict(counts)}


def create_buyer_workbook(source_path: Path, target_path: Path, base_headers: list[str], rows: list[dict]) -> dict:
    wb = openpyxl.load_workbook(source_path)
    for name in list(wb.sheetnames):
        if name.startswith("购方-") or name in {"购方索引", "统计口径"}:
            del wb[name]
    index = wb.create_sheet("购方索引")
    notes = wb.create_sheet("统计口径")
    index.append(["购方全名", "工作表名", "记录数", "计入建议合计数", "待人工确认数"])
    notes.append(["口径", "说明"])
    notes.append(["全部记录", "保留所有原始记录，不自动删除任何发票。"])
    notes.append(["建议合计", "仅排除 SHA256 完全相同文件的后续副本；号码重复但内容不同、号码缺失均保留并提示人工确认。"])
    notes.append(["最终入账", "必须由财务人工确认，本工具不作税务或入账决定。"])
    for sheet in (index, notes):
        for cell in sheet[1]:
            cell.fill, cell.font = HEADER_FILL, HEADER_FONT
    groups = defaultdict(list)
    for row in rows:
        groups[str(row["购方分类"])].append(row)
    used = set(wb.sheetnames)
    summary = {}
    extra = ["购方分类来源", "重复状态", "计入建议合计"]
    output_headers = base_headers + extra
    for buyer, group in sorted(groups.items(), key=lambda item: (-len(item[1]), item[0])):
        sheet_name = unique_sheet_name(buyer, used)
        sheet = wb.create_sheet(sheet_name)
        sheet.append(output_headers)
        for cell in sheet[1]:
            cell.fill, cell.font, cell.alignment = HEADER_FILL, HEADER_FONT, Alignment(horizontal="center")
        for row in group:
            sheet.append([row.get(h) for h in base_headers] + [row.get(h) for h in extra])
            r = sheet.max_row
            fill = DUP_FILL if row["计入建议合计"] == "否" else NOTE_FILL if "确认" in row["重复状态"] or "无法" in row["重复状态"] else None
            for cell in sheet[r]:
                cell.border = BORDER
                if fill:
                    cell.fill = fill
        all_rows = [r for r in group if isinstance(r.get("价税合计"), (int, float))]
        suggested = [r for r in all_rows if r["计入建议合计"] == "是"]
        def totals(items):
            return [round(sum(float(r.get(k) or 0) for r in items), 2) for k in ("价税合计", "合计金额", "合计税额")]
        sheet.append(["合计（全部记录）", "", "", ""] + totals(all_rows))
        for cell in sheet[sheet.max_row]:
            cell.fill, cell.font = TOTAL_FILL, Font(bold=True)
        sheet.append(["建议合计（仅排除文件完全重复副本）", "", "", ""] + totals(suggested))
        for cell in sheet[sheet.max_row]:
            cell.fill, cell.font = TOTAL_FILL, Font(bold=True)
        review_count = sum("确认" in r["重复状态"] or "无法" in r["重复状态"] or r["购方分类"] == "未分类" for r in group)
        index.append([buyer, sheet_name, len(group), len(suggested), review_count])
        summary[buyer] = {"sheet": sheet_name, "records": len(group), "suggested_records": len(suggested), "manual_review": review_count}
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(output_headers))}{1 + len(group)}"
        for c in range(1, len(output_headers) + 1):
            longest = max(len(str(sheet.cell(r, c).value or "")) for r in range(1, sheet.max_row + 1))
            sheet.column_dimensions[get_column_letter(c)].width = min(max(12, longest + 2), 45)
    target_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(target_path)
    return {"buyers": len(groups), "buyer_sheets": summary, "output_file": str(target_path)}


def organize(workbook: Path, attachments: Path, output_dir: Path, rules_path: Path | None = None, organize_files: bool = True, buyer_sheets: bool = True) -> dict:
    base_headers, raw_rows = rows_from_workbook(workbook)
    rules = load_rules(rules_path)
    rows = enrich_rows(raw_rows, rules)
    result = {"records": len(rows), "unclassified": sum(r["购方分类"] == "未分类" for r in rows), "manual_review": sum("确认" in r["重复状态"] or "无法" in r["重复状态"] for r in rows), "excluded_exact_duplicates": sum(r["计入建议合计"] == "否" for r in rows)}
    if organize_files:
        result["file_organization"] = copy_by_buyer(rows, output_dir / "by_company")
    if buyer_sheets:
        target = output_dir / (workbook.stem + "_购方分表.xlsx")
        result["buyer_workbook"] = create_buyer_workbook(workbook, target, base_headers, rows)
    (output_dir / "buyer-summary.json").write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    return result


def main() -> int:
    parser = argparse.ArgumentParser(description="按购方归档发票原件并生成安全的购方分表")
    parser.add_argument("--workbook", type=Path, required=True, help="基础发票登记表")
    parser.add_argument("--attachments", type=Path, help="附件目录；默认使用登记表中的本地文件路径")
    parser.add_argument("--output-dir", type=Path, required=True)
    parser.add_argument("--rules", type=Path, help="用户私有购方映射 JSON")
    parser.add_argument("--no-organize-files", action="store_true")
    parser.add_argument("--no-buyer-sheets", action="store_true")
    args = parser.parse_args()
    result = organize(args.workbook, args.attachments or args.output_dir / "attachments", args.output_dir, args.rules, not args.no_organize_files, not args.no_buyer_sheets)
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
